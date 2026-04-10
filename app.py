from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import openpyxl
import datetime
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'  # Required for session

EXCEL_FILE = "exam_results.xlsx"

#  ---------- PYTHON QUESTIONS (original) ----------
QUESTIONS = [
    {
        "question": "1. What is the correct way to create a list in Python?",
        "options": ["A. list = (1, 2, 3)", "B. list = [1, 2, 3]", "C. list = {1, 2, 3}", "D. list = <1, 2, 3>"],
        "answer": "B"
    },
    {
        "question": "2. Which of the following is used to define a function in Python?",
        "options": ["A. define", "B. function", "C. def", "D. func"],
        "answer": "C"
    },
    {
        "question": "3. What is the output of print(2 ** 3)?",
        "options": ["A. 6", "B. 8", "C. 9", "D. 5"],
        "answer": "B"
    },
    {
        "question": "4. Which keyword is used to import a module in Python?",
        "options": ["A. include", "B. using", "C. import", "D. require"],
        "answer": "C"
    },
    {
        "question": "5. What does the 'len()' function do?",
        "options": ["A. Returns the length of an object", "B. Converts to lowercase", "C. Rounds a number", "D. Finds the maximum value"],
        "answer": "A"
    }
]

# ------------------------- EXCEL HELPER -------------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Results"
        headers = ["#", "Student Name", "Class", "Registration No.",
                   "Score", "Total", "Percentage (%)", "Status", "Date & Time", "Remarks"]
        ws.append(headers)
        wb.save(EXCEL_FILE)

def save_result(name, cls, reg, score, total, remarks=""):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    row_num = ws.max_row + 1
    pct = round((score / total) * 100, 1)
    status = "PASS" if pct >= 50 else "FAIL"
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([row_num - 1, name, cls, reg, score, total, f"{pct}%", status, dt, remarks])
    wb.save(EXCEL_FILE)

# ------------------------- ROUTES -------------------------
@app.route('/')
def index():
    return render_template('register.html')

@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    name = data.get('name', '').strip()
    cls = data.get('cls', '').strip()
    reg = data.get('reg', '').strip()
    if not all([name, cls, reg]):
        return jsonify({"error": "All fields required"}), 400
    session['student'] = {'name': name, 'cls': cls, 'reg': reg}
    return jsonify({"success": True})

@app.route('/exam')
def exam():
    if 'student' not in session:
        return redirect(url_for('index'))
    return render_template('exam.html', questions=QUESTIONS)

@app.route('/submit_exam', methods=['POST'])
def submit_exam():
    if 'student' not in session:
        return jsonify({"error": "No session"}), 401
    data = request.get_json()
    answers = data.get('answers', {})
    timeout = data.get('timeout', False)
    student = session['student']
    # Calculate score
    score = 0
    for idx, q in enumerate(QUESTIONS):
        if str(idx) in answers and answers[str(idx)] == q['answer']:
            score += 1
    total = len(QUESTIONS)
    reason = "Time Expired" if timeout else "Student Submitted"
    save_result(student['name'], student['cls'], student['reg'], score, total, reason)
    session.clear()
    return jsonify({"score": score, "total": total, "passed": score >= total/2})

@app.route('/session_info')
def session_info():
    student = session.get('student')
    if student:
        return jsonify(student)
    return jsonify({}), 401

@app.route('/result')
def result():
    score = request.args.get('score', type=int)
    total = request.args.get('total', type=int)
    passed = request.args.get('passed') == 'true'
    timeout = request.args.get('timeout') == 'true'
    return render_template('result.html', score=score, total=total, passed=passed, timeout=timeout)

if __name__ == '__main__':
    app.run(debug=True)